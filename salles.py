import os
import sqlite3
import time
from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import Qt, pyqtSignal
from datetime import datetime
import pandas as pd
import random
import unicodedata
from reportlab.lib.pagesizes import A4
from reportlab import *

class SallesEntryDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Entrer les Salles")
        self.setMinimumWidth(800)
        self.setStyleSheet("""
            QDialog {
                background-color: rgba(40, 40, 50, 0.7);
                border-radius: 10px;
                border: 1px solid rgba(255,255,255,0.1);
            }
        """)
        layout = QVBoxLayout(self)
        title = QLabel("<span style='color:#3498db; font-size:22px; font-weight:bold; font-family:Arial;'>Entrer les Salles et leurs Capacités</span>")
        title.setTextFormat(Qt.TextFormat.RichText)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Locaux d'examen", "Capacité", "Type", "Climatisé", "Camera"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: rgba(50, 50, 60, 0.7);
                color: white;
                border-radius: 5px;
                font-size: 13px;
                gridline-color: #394150;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                border: none;
                padding: 6px;
            }
        """)
        layout.addWidget(self.table)
        btns_layout = QHBoxLayout()
        btn_add = QPushButton("Ajouter une salle")
        btn_add.setStyleSheet("background-color: #2980b9; color: white; border-radius: 8px; padding: 8px 18px; font-weight: bold;")
        btn_add.clicked.connect(self.add_row)
        btn_del = QPushButton("Supprimer la salle")
        btn_del.setStyleSheet("background-color: #e74c3c; color: white; border-radius: 8px; padding: 8px 18px; font-weight: bold;")
        btn_del.clicked.connect(self.delete_row)
        btns_layout.addWidget(btn_add)
        btns_layout.addWidget(btn_del)
        layout.addLayout(btns_layout)
        btns2_layout = QHBoxLayout()
        btn_ok = QPushButton("Valider")
        btn_ok.setStyleSheet("background-color: #27ae60; color: white; border-radius: 8px; padding: 8px 32px; font-weight: bold;")
        btn_ok.clicked.connect(self.validate_and_accept)
        btn_cancel = QPushButton("Annuler")
        btn_cancel.setStyleSheet("background-color: #7f8c8d; color: white; border-radius: 8px; padding: 8px 32px; font-weight: bold;")
        btn_cancel.clicked.connect(self.reject)
        btns2_layout.addWidget(btn_ok)
        btns2_layout.addWidget(btn_cancel)
        layout.addLayout(btns2_layout)
        self.add_row()  # Start with one row

    def validate_and_accept(self):
        import re
        for row in range(self.table.rowCount()):
            nom = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
            capacite = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
            # Validation du nom (lettres et/ou chiffres uniquement)
            if nom and not re.match(r'^[A-Za-z0-9\s]+$', nom):
                QMessageBox.warning(self, "Erreur de saisie", f"Le nom de la salle '{nom}' n'est pas valide. Utilisez uniquement des lettres et/ou des chiffres.")
                return
            # Validation de la capacité (nombre entier positif)
            if capacite:
                try:
                    cap = int(capacite)
                    if cap <= 0:
                        raise ValueError
                except Exception:
                    QMessageBox.warning(self, "Erreur de saisie", f"La capacité de la salle '{nom}' doit être un nombre entier positif.")
                    return
        self.accept()

    def add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Nom de la salle
        nom_item = QTableWidgetItem("")
        self.table.setItem(row, 0, nom_item)
        
        # Capacité
        capacite_item = QTableWidgetItem("")
        self.table.setItem(row, 1, capacite_item)
        
        # Type (combobox)
        combo_type = QComboBox()
        combo_type.addItems(["Grande", "Petite"])
        self.table.setCellWidget(row, 2, combo_type)
        
        # Climatisé (combobox)
        combo_clim = QComboBox()
        combo_clim.addItems(["Oui", "Non"])
        self.table.setCellWidget(row, 3, combo_clim)
        
        # Caméra (combobox)
        combo_cam = QComboBox()
        combo_cam.addItems(["Oui", "Non"])
        self.table.setCellWidget(row, 4, combo_cam)

    def delete_row(self):
        row = self.table.currentRow()
        if row >= 0:
            self.table.removeRow(row)

    def get_data(self):
        data = []
        for row in range(self.table.rowCount()):
            nom = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
            capacite = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
            type_salle = self.table.cellWidget(row, 2).currentText()
            clim = self.table.cellWidget(row, 3).currentText()
            cam = self.table.cellWidget(row, 4).currentText()
            if nom and capacite:  # Ne prendre que les salles valides
                data.append({
                    "nom": nom,
                    "capacite": capacite,
                    "type": type_salle,
                    "climatise": clim,
                    "camera": cam
                })
        return data

class CentreDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ajouter un Centre d'Examen")
        self.setMinimumWidth(600)
        self.setStyleSheet("""
            QDialog {
                background-color: rgba(40, 40, 50, 0.7);
                border-radius: 10px;
                border: 1px solid rgba(255,255,255,0.1);
            }
        """)
        
        layout = QVBoxLayout(self)
        
        # Titre
        title = QLabel("<span style='color:#3498db; font-size:22px; font-weight:bold; font-family:Arial;'>Nouveau Centre d'Examen</span>")
        title.setTextFormat(Qt.TextFormat.RichText)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Nom du centre
        self.nom_centre = QLineEdit()
        self.nom_centre.setPlaceholderText("Nom du centre d'examen ")
        self.nom_centre.setStyleSheet("""
            QLineEdit {
                font-size: 16px;
                padding: 8px;
                border-radius: 5px;
                background: rgba(60, 60, 70, 0.7);
                color: white;
                border: 1px solid rgba(255,255,255,0.1);
            }
        """)
        layout.addWidget(QLabel("Nom du centre:"))
        layout.addWidget(self.nom_centre)
        
        # Bouton pour ajouter des salles
        btn_ajouter_salles = QPushButton("Ajouter les salles pour ce centre")
        btn_ajouter_salles.setStyleSheet("""
            QPushButton {
                background-color: #2980b9;
                color: white;
                border-radius: 8px;
                padding: 10px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1d6fa5;
            }
        """)
        btn_ajouter_salles.clicked.connect(self.ajouter_salles)
        layout.addWidget(btn_ajouter_salles)
        
        # Boutons Valider/Annuler
        btn_layout = QHBoxLayout()
        btn_valider = QPushButton("Valider")
        btn_valider.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #219955;
            }
        """)
        btn_valider.clicked.connect(self.accept)
        
        btn_annuler = QPushButton("Annuler")
        btn_annuler.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        btn_annuler.clicked.connect(self.reject)
        
        btn_layout.addWidget(btn_valider)
        btn_layout.addWidget(btn_annuler)
        layout.addLayout(btn_layout)
        
        self.salles_data = []

    def ajouter_salles(self):
        dialog = SallesEntryDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.salles_data = dialog.get_data()

    def get_data(self):
        return {
            "nom_centre": self.nom_centre.text().strip(),
            "salles": self.salles_data
        }

class CentresSallesEntryDialog(QDialog):
    data_changed = pyqtSignal()  # Définir le signal
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.df_imported = None  # Pour stocker les données importées d'origine
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("Gestion des Centres et Salles")
        self.setMinimumWidth(1000)
        self.setMinimumHeight(600)
        self.setStyleSheet("""
            QDialog {
                background-color: rgba(40, 40, 50, 0.7);
                border-radius: 10px;
                border: 1px solid rgba(255,255,255,0.1);
            }
        """)
        self.layout = QVBoxLayout(self)
        # Titre
        title = QLabel("<span style='color:#3498db; font-size:22px; font-weight:bold; font-family:Arial;'>Gestion des Centres d'Examen et de leurs Salles</span>")
        title.setTextFormat(Qt.TextFormat.RichText)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(title)
        # Tableau des centres et salles (sans colonne Actions)
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Centre", "Locaux d'examen", "Capacité", "Climatisé", "Caméra", "Type"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: rgba(50, 50, 60, 0.7);
                color: white;
                border-radius: 5px;
                font-size: 13px;
                gridline-color: #394150;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                border: none;
                padding: 6px;
            }
        """)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.layout.addWidget(self.table)
        # Boutons
        btn_layout = QHBoxLayout()
        btn_ajouter_centre = QPushButton("Ajouter un Centre")
        btn_ajouter_centre.setStyleSheet("""
            QPushButton {
                background-color: #2980b9;
                color: white;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1d6fa5;
            }
        """)
        btn_ajouter_centre.clicked.connect(self.ajouter_centre)

        # Bouton pour importer depuis Excel
        btn_importer = QPushButton("Importer depuis Excel")
        btn_importer.setStyleSheet("""
            QPushButton {
                background-color: #8e44ad;
                color: white;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #6c3483;
            }
        """)
        btn_importer.clicked.connect(self.importer_excel)

        btn_afficher_totaux = QPushButton("Afficher les Totaux")
        btn_afficher_totaux.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #219955;
            }
        """)
        btn_afficher_totaux.clicked.connect(self.afficher_totaux)
        # Nouveau bouton pour gérer les salles du centre sélectionné
        btn_gerer_salles = QPushButton("Gérer les salles du centre")
        btn_gerer_salles.setStyleSheet("""
            QPushButton {
                background-color: #f1c40f;
                color: #222;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #f39c12;
            }
        """)
        btn_gerer_salles.clicked.connect(self.gerer_salles_centre_selectionne)
        btn_valider = QPushButton("Valider")
        btn_valider.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        btn_valider.clicked.connect(self.accept)
        btn_annuler = QPushButton("Annuler")
        btn_annuler.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        btn_annuler.clicked.connect(self.reject)
        btn_layout.addWidget(btn_ajouter_centre)
        btn_layout.addWidget(btn_importer)
        btn_layout.addWidget(btn_afficher_totaux)
        btn_layout.addWidget(btn_gerer_salles)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_valider)
        btn_layout.addWidget(btn_annuler)
        self.layout.addLayout(btn_layout)
        self.centres = []
        # Charger les données existantes
        self.charger_donnees()

    def charger_donnees(self):
        """Charge les données depuis la base de données"""
        try:
            # Initialiser les centres
            self.centres = []
            
            # Charger les données
            centres_charges = charger_centres_et_salles()
            if centres_charges:
                self.centres = centres_charges
                
            # Mettre à jour l'interface
            self.mettre_a_jour_tableau()
            
        except sqlite3.OperationalError as e:
            QMessageBox.warning(self, "Base de données occupée", 
                              "La base de données est temporairement inaccessible. Veuillez réessayer.")
        except Exception as e:
            QMessageBox.critical(self, "Erreur de chargement", 
                               f"Erreur lors du chargement des données : {str(e)}")

    def sauvegarder_donnees(self):
        """Sauvegarde les données dans la base de données"""
        try:
            if not self.centres:
                self.centres = []
            time.sleep(0.5)  # Attendre un court instant avant de sauvegarder
            sauvegarder_centres_et_salles(self.centres)
            self.data_changed.emit()
            return True
        except Exception as e:
            QMessageBox.critical(self, "Erreur de sauvegarde", 
                               f"Erreur lors de la sauvegarde : {e}")
            return False

    def ajouter_centre(self):
        dialog = CentreDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            if data["nom_centre"] and data["salles"]:
                # Vérifier si le centre existe déjà
                centre_existant = None
                for centre in self.centres:
                    if centre["nom_centre"] == data["nom_centre"]:
                        centre_existant = centre
                        break
                
                if centre_existant:
                    # Ajouter les nouvelles salles au centre existant
                    centre_existant["salles"].extend(data["salles"])
                else:
                    # Ajouter le nouveau centre
                    self.centres.append(data)
                
                # Sauvegarder et mettre à jour
                self.sauvegarder_donnees()
                self.mettre_a_jour_tableau()

    def mettre_a_jour_tableau(self):
        try:
            self.table.setRowCount(0)  # Vider le tableau
            centres_uniques = {}  # Dictionnaire pour stocker les centres uniques
            
            # Charger les données fraîches depuis la base de données
            self.centres = charger_centres_et_salles()
            
            # Regrouper les salles par centre
            for centre in self.centres:
                nom_centre = centre["nom_centre"]
                if nom_centre and nom_centre not in centres_uniques:
                    centres_uniques[nom_centre] = {
                        "nom_centre": nom_centre,
                        "salles": []
                    }
                if nom_centre:
                    centres_uniques[nom_centre]["salles"].extend(centre["salles"])
            
            # Mettre à jour le tableau avec les centres uniques
            for centre in centres_uniques.values():
                start_row = self.table.rowCount()
                for i, salle in enumerate(centre["salles"]):
                    row = self.table.rowCount()
                    self.table.insertRow(row)
                    
                    # Centre (uniquement sur la première ligne)
                    if i == 0:
                        item = QTableWidgetItem(centre["nom_centre"])
                        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        self.table.setItem(row, 0, item)
                    else:
                        self.table.setItem(row, 0, QTableWidgetItem(""))
                    
                    # Informations sur la salle
                    salle_item = QTableWidgetItem(salle["nom"])
                    if salle["type"] == "Petite":
                        salle_item.setForeground(QColor("#FF0000"))  # Rouge pour les petites salles
                    self.table.setItem(row, 1, salle_item)
                    
                    capacite_item = QTableWidgetItem(str(salle["capacite"]))
                    capacite_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(row, 2, capacite_item)
                    
                    clim_item = QTableWidgetItem(salle["climatise"] if salle["climatise"] else "-")
                    clim_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(row, 3, clim_item)
                    
                    cam_item = QTableWidgetItem(salle["camera"] if salle["camera"] else "-")
                    cam_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(row, 4, cam_item)
                    
                    type_item = QTableWidgetItem(salle["type"] if salle["type"] else "-")
                    type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(row, 5, type_item)
                
                if len(centre["salles"]) > 1:
                    self.table.setSpan(start_row, 0, len(centre["salles"]), 1)
            
            # Mettre à jour self.centres avec les centres uniques
            self.centres = list(centres_uniques.values())
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la mise à jour du tableau : {str(e)}")

    def gerer_salles_centre_selectionne(self):
        if not self.centres:
            QMessageBox.warning(self, "Aucun centre", "Aucun centre n'a été saisi. Veuillez d'abord ajouter un centre.")
            return
        # Récupérer le centre sélectionné (ligne sélectionnée)
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Aucun centre sélectionné", "Veuillez sélectionner une ligne du centre à gérer.")
            return
        # Trouver le nom du centre (colonne 0, première ligne fusionnée)
        nom_centre = self.table.item(row, 0).text()
        if not nom_centre:
            for r in range(row, -1, -1):
                nom_centre = self.table.item(r, 0).text()
                if nom_centre:
                    break
        for centre in self.centres:
            if centre["nom_centre"] == nom_centre:
                # Ouvrir un QDialog pour modifier le nom du centre et ses salles
                dialog = QDialog(self)
                dialog.setWindowTitle(f"Gérer le centre et ses salles : {nom_centre}")
                dialog.setMinimumWidth(700)
                dialog.setStyleSheet("""
                    QDialog {
                        background-color: rgba(40, 40, 50, 0.7);
                        border-radius: 10px;
                        border: 1px solid rgba(255,255,255,0.1);
                    }
                """)
                layout = QVBoxLayout(dialog)
                # Champ pour modifier le nom du centre
                lbl_nom = QLabel("Nom du centre :")
                edit_nom = QLineEdit(centre["nom_centre"])
                edit_nom.setStyleSheet("""
                    QLineEdit {
                        font-size: 16px;
                        padding: 8px;
                        border-radius: 5px;
                        background: rgba(60, 60, 70, 0.7);
                        color: white;
                        border: 1px solid rgba(255,255,255,0.1);
                    }
                """)
                layout.addWidget(lbl_nom)
                layout.addWidget(edit_nom)
                # Table des salles (SallesEntryDialog simplifié)
                table = QTableWidget(0, 5)
                table.setHorizontalHeaderLabels(["Nom de la salle", "Capacité", "Type", "Climatisé", "Caméra"])
                table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
                table.setStyleSheet("""
                    QTableWidget {
                        background-color: rgba(50, 50, 60, 0.7);
                        color: white;
                        border-radius: 5px;
                        font-size: 13px;
                        gridline-color: #394150;
                    }
                    QHeaderView::section {
                        background-color: #3498db;
                        color: white;
                        font-weight: bold;
                        border: none;
                        padding: 6px;
                    }
                """)
                for salle in centre["salles"]:
                    row_idx = table.rowCount()
                    table.insertRow(row_idx)
                    table.setItem(row_idx, 0, QTableWidgetItem(salle["nom"]))
                    table.setItem(row_idx, 1, QTableWidgetItem(str(salle["capacite"])))
                    combo_type = QComboBox()
                    combo_type.addItems(["Grande", "Petite"])
                    combo_type.setCurrentText(salle.get("type", "Grande") if salle.get("type") in ["Grande", "Petite"] else "Grande")
                    table.setCellWidget(row_idx, 2, combo_type)
                    combo_clim = QComboBox()
                    combo_clim.addItems(["Oui", "Non"])
                    combo_clim.setCurrentText(salle["climatise"])
                    table.setCellWidget(row_idx, 3, combo_clim)
                    combo_cam = QComboBox()
                    combo_cam.addItems(["Oui", "Non"])
                    combo_cam.setCurrentText(salle["camera"])
                    table.setCellWidget(row_idx, 4, combo_cam)
                layout.addWidget(table)
                # Boutons pour ajouter/supprimer une salle
                btns_layout = QHBoxLayout()
                btn_add = QPushButton("Ajouter une salle")
                btn_add.setStyleSheet("background-color: #2980b9; color: white; border-radius: 8px; padding: 8px 18px; font-weight: bold;")
                btn_add.clicked.connect(lambda: self._add_row_to_table(table))
                btn_del = QPushButton("Supprimer la salle")
                btn_del.setStyleSheet("background-color: #e74c3c; color: white; border-radius: 8px; padding: 8px 18px; font-weight: bold;")
                btn_del.clicked.connect(lambda: self._delete_row_from_table(table))
                btns_layout.addWidget(btn_add)
                btns_layout.addWidget(btn_del)
                layout.addLayout(btns_layout)
                # Boutons Valider/Annuler
                btns2_layout = QHBoxLayout()
                btn_ok = QPushButton("Valider")
                btn_ok.setStyleSheet("background-color: #27ae60; color: white; border-radius: 8px; padding: 8px 32px; font-weight: bold;")
                btn_ok.clicked.connect(dialog.accept)
                btn_cancel = QPushButton("Annuler")
                btn_cancel.setStyleSheet("background-color: #7f8c8d; color: white; border-radius: 8px; padding: 8px 32px; font-weight: bold;")
                btn_cancel.clicked.connect(dialog.reject)
                btns2_layout.addWidget(btn_ok)
                btns2_layout.addWidget(btn_cancel)
                layout.addLayout(btns2_layout)
                if dialog.exec() == QDialog.DialogCode.Accepted:
                    # Appliquer les modifications
                    centre["nom_centre"] = edit_nom.text().strip()
                    new_salles = []
                    for r in range(table.rowCount()):
                        nom = table.item(r, 0).text() if table.item(r, 0) else ""
                        capacite = table.item(r, 1).text() if table.item(r, 1) else ""
                        type_salle = table.cellWidget(r, 2).currentText()
                        clim = table.cellWidget(r, 3).currentText()
                        cam = table.cellWidget(r, 4).currentText()
                        if nom and capacite:
                            new_salles.append({
                                "nom": nom,
                                "capacite": capacite,
                                "type": type_salle,
                                "climatise": clim,
                                "camera": cam
                            })
                    centre["salles"] = new_salles
                    self.sauvegarder_donnees()
                    self.mettre_a_jour_tableau()
                break

    def _add_row_to_table(self, table):
        row = table.rowCount()
        table.insertRow(row)
        table.setItem(row, 0, QTableWidgetItem(""))
        table.setItem(row, 1, QTableWidgetItem(""))
        combo_type = QComboBox()
        combo_type.addItems(["Grande", "Petite"])
        table.setCellWidget(row, 2, combo_type)
        combo_clim = QComboBox()
        combo_clim.addItems(["Oui", "Non"])
        table.setCellWidget(row, 3, combo_clim)
        combo_cam = QComboBox()
        combo_cam.addItems(["Oui", "Non"])
        table.setCellWidget(row, 4, combo_cam)

    def _delete_row_from_table(self, table):
        row = table.currentRow()
        if row >= 0:
            table.removeRow(row)

    def afficher_totaux(self):
        totaux = {}
        for centre in self.centres:
            total = 0
            total_grande = 0
            for salle in centre["salles"]:
                try:
                    capacite = int(salle["capacite"])
                    total += capacite
                    if salle["type"] == "Grande":
                        total_grande += capacite
                except ValueError:
                    pass
            totaux[centre["nom_centre"]] = {
                "total": total,
                "total_grande": total_grande
            }
        
        # Afficher les totaux dans une boîte de dialogue
        dialog = QDialog(self)
        dialog.setWindowTitle("Totaux par Centre")
        dialog.setMinimumWidth(500)
        dialog.setStyleSheet("""
            QDialog {
                background-color: rgba(40, 40, 50, 0.7);
                border-radius: 10px;
                border: 1px solid rgba(255,255,255,0.1);
            }
        """)
        
        layout = QVBoxLayout(dialog)
        
        title = QLabel("<span style='color:#3498db; font-size:18px; font-weight:bold;'>Totaux par Centre d'Examen</span>")
        title.setTextFormat(Qt.TextFormat.RichText)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        table = QTableWidget(len(totaux), 3)
        table.setHorizontalHeaderLabels(["Centre", "Total Places", "Total (Grandes Salles)"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setStyleSheet("""
            QTableWidget {
                background-color: rgba(50, 50, 60, 0.7);
                color: white;
                border-radius: 5px;
                font-size: 13px;
                gridline-color: #394150;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                border: none;
                padding: 6px;
            }
        """)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        for i, (centre, data) in enumerate(totaux.items()):
            table.setItem(i, 0, QTableWidgetItem(centre))
            table.setItem(i, 1, QTableWidgetItem(str(data["total"])))
            table.setItem(i, 2, QTableWidgetItem(str(data["total_grande"])))
        
        layout.addWidget(table)
        
        btn_fermer = QPushButton("Fermer")
        btn_fermer.setStyleSheet("""
            QPushButton {
                background-color: #2980b9;
                color: white;
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1d6fa5;
            }
        """)
        btn_fermer.clicked.connect(dialog.accept)
        layout.addWidget(btn_fermer, alignment=Qt.AlignmentFlag.AlignCenter)
        
        dialog.exec()

    def get_data(self):
        data = []
        for centre in self.centres:
            for salle in centre["salles"]:
                data.append({
                    "centre": centre["nom_centre"],
                    "nom": salle["nom"],
                    "capacite": salle["capacite"],
                    "climatise": salle["climatise"],
                    "camera": salle["camera"],
                    "type": salle["type"]
                })
        return data

    def ajouter_salle_au_centre(self, nom_centre):
        # Ouvre un mini-dialogue pour ajouter une salle à un centre existant
        dialog = SallesEntryDialog(self)
        dialog.setWindowTitle(f"Ajouter une salle au centre {nom_centre}")
        dialog.table.setRowCount(0)  # Vider le tableau
        dialog.add_row()
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_salles = dialog.get_data()
            for centre in self.centres:
                if centre["nom_centre"] == nom_centre:
                    centre["salles"].extend(new_salles)
                    self.sauvegarder_donnees()
                    break
            self.mettre_a_jour_tableau()
    def importer_excel(self):
        import pandas as pd
        import unicodedata
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import openpyxl
        file, _ = QFileDialog.getOpenFileName(self, "Choisir un fichier Excel", "", "Fichiers Excel (*.xlsx *.xls)")
        if not file:
            return
        try:
            df = pd.read_excel(file)
            # Normaliser les noms de colonnes (enlever accents, espaces, tout en minuscule)
            def normalize(col):
                col = str(col)
                col = ''.join(c for c in unicodedata.normalize('NFD', col) if unicodedata.category(c) != 'Mn')
                return col.strip().lower().replace(' ', '')
            df.columns = [normalize(c) for c in df.columns]
            # Trouver les bons noms de colonnes
            col_centre = next((c for c in df.columns if 'centre' in c), None)
            col_salle = next((c for c in df.columns if 'locaux' in c or 'salle' in c), None)
            col_capacite = next((c for c in df.columns if 'capacite' in c), None)
            col_clim = next((c for c in df.columns if 'clim' in c), None)
            col_camera = next((c for c in df.columns if 'camera' in c), None)
            if not (col_centre and col_salle and col_capacite):
                QMessageBox.warning(self, "Colonnes manquantes", "Le fichier doit contenir au moins les colonnes centre, salle et capacité.")
                return
            # Charger le workbook openpyxl pour lire la couleur
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            # Associer index pandas et lignes Excel (header=0)
            excel_headers = [normalize(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            idx_salle = excel_headers.index(col_salle) if col_salle in excel_headers else None
            centres_dict = {}
            last_centre = None
            for idx, row in df.iterrows():
                centre = str(row.get(col_centre, "")).strip()
                if centre and centre.lower() != 'nan':
                    last_centre = centre
                if not last_centre:
                    continue
                if last_centre not in centres_dict:
                    centres_dict[last_centre] = []
                salle = str(row.get(col_salle, "")).strip()
                capacite = row.get(col_capacite, "")
                climatise = str(row.get(col_clim, "")).strip() if col_clim else ''
                camera = str(row.get(col_camera, "")).strip() if col_camera else ''
                # Filtrer les lignes vides, totaux, ou non numériques
                if not salle or salle.lower().startswith('total') or salle == 'nan':
                    continue
                try:
                    cap_int = int(str(capacite).strip())
                except:
                    continue
                # Vérifier la couleur de la cellule du nom de salle (ligne Excel = idx+2 car header=1)
                type_salle = "Grande"
                if idx_salle is not None:
                    cell = ws.cell(row=idx+2, column=idx_salle+1)
                    color = None
                    if cell.font and cell.font.color:
                        if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                            color = str(cell.font.color.rgb)
                        elif hasattr(cell.font.color, 'type') and cell.font.color.type == 'rgb' and hasattr(cell.font.color, 'value'):
                            color = str(cell.font.color.value)
                    if color:
                        color_str = color.upper()
                        # Accepte FFFF0000, FF0000, 00FF0000, etc.
                        if color_str.endswith('FF0000') or color_str == 'FFFF0000' or color_str == 'FF0000':
                            type_salle = "Petite"
                centres_dict[last_centre].append({
                    "nom": salle,
                    "capacite": cap_int,
                    "type": type_salle,
                    "climatise": climatise,
                    "camera": camera
                })
            centres = [
                {"nom_centre": nom, "salles": salles}
                for nom, salles in centres_dict.items() if salles
            ]
            if not centres:
                QMessageBox.warning(self, "Aucun centre", "Aucun centre ou salle valide trouvé dans le fichier.")
                return

            try:
                # Initialiser la base de données
                os.makedirs('database', exist_ok=True)
                db_path = os.path.join('database', 'salles.db')
                
                # Supprimer l'ancien fichier de base de données s'il existe
                if os.path.exists(db_path):
                    os.remove(db_path)
                
                # Créer une nouvelle connexion
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                
                # Créer les nouvelles tables avec les bonnes contraintes
                cursor.execute('''CREATE TABLE centres
                                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                 nom TEXT UNIQUE NOT NULL)''')
                
                cursor.execute('''CREATE TABLE salles
                                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                 centre_id INTEGER NOT NULL,
                                 nom TEXT NOT NULL,
                                 capacite INTEGER NOT NULL,
                                 type TEXT NOT NULL,
                                 climatise TEXT NOT NULL,
                                 camera TEXT NOT NULL,
                                 FOREIGN KEY (centre_id) REFERENCES centres(id),
                                 UNIQUE(centre_id, nom))''')
                
                # Insérer les nouvelles données
                for centre in centres:
                    cursor.execute("INSERT INTO centres (nom) VALUES (?)",
                                 (centre["nom_centre"],))
                    centre_id = cursor.lastrowid
                    
                    for salle in centre["salles"]:
                        cursor.execute("""INSERT INTO salles 
                                     (centre_id, nom, capacite, type, climatise, camera)
                                     VALUES (?, ?, ?, ?, ?, ?)""",
                                     (centre_id, salle["nom"], salle["capacite"],
                                      salle["type"], salle["climatise"], salle["camera"]))
                
                conn.commit()
                
                # Mettre à jour l'interface et sauvegarder dans la base de données
                self.centres = centres
                time.sleep(0.5)  # Attendre un court instant avant de sauvegarder
                self.sauvegarder_donnees()  # Sauvegarder explicitement après l'importation
                self.mettre_a_jour_tableau()
                self.last_imported_file = os.path.basename(file)
                QMessageBox.information(self, "Importation réussie", 
                                     "Les données ont été importées avec succès et remplacent les anciennes données.")
                
            except Exception as e:
                QMessageBox.critical(self, "Erreur d'importation", 
                                   f"Erreur lors de l'importation : {e}")
            finally:
                if 'conn' in locals():
                    conn.close()
            QMessageBox.information(self, "Importation réussie", "Les centres et salles ont été importés avec succès.")
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'importation", f"Erreur lors de l'importation : {e}")

def sauvegarder_centres_et_salles(centres):
    conn = None
    try:
        # S'assurer que le dossier database existe
        os.makedirs('database', exist_ok=True)
        db_path = os.path.join('database', 'salles.db')
        
        # Établir la connexion avec un timeout
        conn = sqlite3.connect(db_path, timeout=20)
        cursor = conn.cursor()
        
        # Créer les tables avec les bonnes contraintes
        cursor.execute('''CREATE TABLE IF NOT EXISTS centres
                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                       nom TEXT UNIQUE NOT NULL)''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS salles
                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                       centre_id INTEGER NOT NULL,
                       nom TEXT NOT NULL,
                       capacite INTEGER NOT NULL,
                       type TEXT NOT NULL,
                       climatise TEXT NOT NULL,
                       camera TEXT NOT NULL,
                       FOREIGN KEY (centre_id) REFERENCES centres(id),
                       UNIQUE(centre_id, nom))''')
        
        # Vider les tables existantes
        cursor.execute("DELETE FROM salles")
        cursor.execute("DELETE FROM centres")
        
        # Insérer les nouvelles données avec gestion des doublons
        for centre in centres:
            # Insérer ou récupérer le centre
            cursor.execute("INSERT OR IGNORE INTO centres (nom) VALUES (?)",
                         (centre["nom_centre"],))
            cursor.execute("SELECT id FROM centres WHERE nom = ?",
                         (centre["nom_centre"],))
            centre_id = cursor.fetchone()[0]
            
            # Insérer les salles en évitant les doublons
            for salle in centre["salles"]:
                cursor.execute("""INSERT OR REPLACE INTO salles 
                                (centre_id, nom, capacite, type, climatise, camera)
                                VALUES (?, ?, ?, ?, ?, ?)""",
                             (centre_id, salle["nom"], salle["capacite"],
                              salle["type"], salle["climatise"], salle["camera"]))
        
        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Erreur lors de la sauvegarde : {e}")
        raise  # Relever l'exception pour le débogage
    finally:
        if conn:
            conn.close()

def charger_centres_et_salles():
    conn = None
    try:
        db_path = os.path.join('database', 'salles.db')
        if not os.path.exists(db_path):
            return []
            
        conn = sqlite3.connect(db_path, timeout=20)
        cursor = conn.cursor()
        
        # Vérifier que les tables existent et les créer si nécessaire
        cursor.execute('''CREATE TABLE IF NOT EXISTS centres
                         (id INTEGER PRIMARY KEY AUTOINCREMENT,
                          nom TEXT UNIQUE NOT NULL)''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS salles
                         (id INTEGER PRIMARY KEY AUTOINCREMENT,
                          centre_id INTEGER NOT NULL,
                          nom TEXT NOT NULL,
                          capacite INTEGER NOT NULL,
                          type TEXT NOT NULL,
                          climatise TEXT NOT NULL,
                          camera TEXT NOT NULL,
                          FOREIGN KEY (centre_id) REFERENCES centres(id),
                          UNIQUE(centre_id, nom))''')
        
        # Charger les centres et leurs salles
        cursor.execute("""SELECT c.nom,
                                s.nom, s.capacite, s.type, s.climatise, s.camera
                         FROM centres c
                         LEFT JOIN salles s ON c.id = s.centre_id
                         ORDER BY c.id, s.id""")
        
        rows = cursor.fetchall()
        centres = {}
        
        for row in rows:
            nom_centre = row[0]
            if row[1]:  # Si la salle existe
                salle = {
                    "nom": row[1],
                    "capacite": row[2],
                    "type": row[3],
                    "climatise": row[4],
                    "camera": row[5]
                }
                if nom_centre in centres:
                    centres[nom_centre]["salles"].append(salle)
                else:
                    centres[nom_centre] = {
                        "nom_centre": nom_centre,
                        "salles": [salle]
                    }
            elif nom_centre not in centres:
                centres[nom_centre] = {
                    "nom_centre": nom_centre,
                    "salles": []
                }
        
        return list(centres.values())
    except Exception as e:
        print(f"Erreur lors du chargement : {e}")
        return []
    finally:
        if conn:
            conn.close()

